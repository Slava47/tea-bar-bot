"""
Веб-сайт аналитики для бара «Ли Бо».
Запуск: python web.py
"""

import csv
import io
import os
import sqlite3

from flask import (
    Flask,
    Response,
    redirect,
    render_template,
    request,
    session,
    url_for,
)

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None

DB_PATH = os.getenv("DB_PATH", "libo.db")
SECRET_KEY = os.getenv("FLASK_SECRET", "super-secret-key-libo-2024")

ADMIN_LOGIN = os.getenv("ADMIN_LOGIN", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "adminLIBO")

app = Flask(__name__)
app.secret_key = SECRET_KEY


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def login_required(f):
    from functools import wraps

    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorated


# ─── Auth ─────────────────────────────────────────────────────────────────────


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        if username == ADMIN_LOGIN and password == ADMIN_PASSWORD:
            session["logged_in"] = True
            return redirect(url_for("dashboard"))
        error = "Неверный логин или пароль"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ─── Dashboard ────────────────────────────────────────────────────────────────


DISPLAY_LIMIT = 100


@app.route("/")
@login_required
def dashboard():
    db = get_db()

    total_users = db.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    total_quizzes = db.execute("SELECT COUNT(*) FROM quiz_sessions").fetchone()[0]
    total_ratings = db.execute("SELECT COUNT(*) FROM cocktail_ratings").fetchone()[0]

    users = db.execute(
        "SELECT u.user_id, u.username, u.first_name, u.last_name, u.created_at, "
        "(SELECT COUNT(*) FROM quiz_sessions q WHERE q.user_id = u.user_id) as quiz_count "
        "FROM users u ORDER BY u.created_at DESC LIMIT ?",
        (DISPLAY_LIMIT,),
    ).fetchall()

    top_cocktails = db.execute(
        "SELECT cocktail_name, COUNT(*) as cnt FROM cocktail_ratings "
        "GROUP BY cocktail_name ORDER BY cnt DESC LIMIT 10"
    ).fetchall()

    db.close()
    return render_template(
        "dashboard.html",
        users=users,
        total_users=total_users,
        total_quizzes=total_quizzes,
        total_ratings=total_ratings,
        top_cocktails=top_cocktails,
        display_limit=DISPLAY_LIMIT,
    )


# ─── User Detail ──────────────────────────────────────────────────────────────


@app.route("/user/<int:user_id>")
@login_required
def user_detail(user_id):
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE user_id = ?", (user_id,)).fetchone()
    if not user:
        return "Пользователь не найден", 404

    total_ratings = db.execute(
        "SELECT COUNT(*) FROM cocktail_ratings WHERE user_id = ?", (user_id,)
    ).fetchone()[0]

    ratings = db.execute(
        "SELECT cocktail_name, rating, review, created_at FROM cocktail_ratings "
        "WHERE user_id = ? ORDER BY created_at DESC LIMIT ?",
        (user_id, DISPLAY_LIMIT),
    ).fetchall()

    quiz_count = db.execute(
        "SELECT COUNT(*) FROM quiz_sessions WHERE user_id = ?", (user_id,)
    ).fetchone()[0]

    db.close()
    return render_template(
        "user_detail.html",
        user=user,
        ratings=ratings,
        quiz_count=quiz_count,
        total_ratings=total_ratings,
        display_limit=DISPLAY_LIMIT,
    )


# ─── Export ───────────────────────────────────────────────────────────────────


@app.route("/export/<fmt>")
@login_required
def export_data(fmt):
    db = get_db()
    users = db.execute(
        "SELECT u.user_id, u.username, u.first_name, u.last_name, u.created_at, "
        "(SELECT COUNT(*) FROM quiz_sessions q WHERE q.user_id = u.user_id) as quiz_count "
        "FROM users u ORDER BY u.created_at DESC"
    ).fetchall()

    ratings = db.execute(
        "SELECT r.user_id, r.cocktail_name, r.rating, r.review, r.created_at "
        "FROM cocktail_ratings r ORDER BY r.created_at DESC"
    ).fetchall()

    quiz_sessions = db.execute(
        "SELECT qs.id, qs.user_id, qs.alcoholic, qs.temperature, qs.taste, "
        "qs.tea_strength, qs.strength, qs.created_at "
        "FROM quiz_sessions qs ORDER BY qs.created_at DESC"
    ).fetchall()

    db.close()

    if fmt == "csv":
        output = io.BytesIO()
        output.write(b'\xef\xbb\xbf')  # UTF-8 BOM for correct Russian characters display
        wrapper = io.TextIOWrapper(output, encoding='utf-8', newline='')
        writer = csv.writer(wrapper)

        writer.writerow(["--- Пользователи ---"])
        writer.writerow(["user_id", "username", "first_name", "last_name", "created_at", "quiz_count"])
        for u in users:
            writer.writerow([u["user_id"], u["username"], u["first_name"], u["last_name"], u["created_at"], u["quiz_count"]])

        writer.writerow([])
        writer.writerow(["--- Оценки ---"])
        writer.writerow(["user_id", "cocktail_name", "rating", "review", "created_at"])
        for r in ratings:
            writer.writerow([r["user_id"], r["cocktail_name"], r["rating"], r["review"], r["created_at"]])

        writer.writerow([])
        writer.writerow(["--- Подборы коктейлей ---"])
        writer.writerow(["id", "user_id", "alcoholic", "temperature", "taste", "tea_strength", "strength", "created_at"])
        for qs in quiz_sessions:
            writer.writerow([qs["id"], qs["user_id"], qs["alcoholic"], qs["temperature"], qs["taste"], qs["tea_strength"], qs["strength"], qs["created_at"]])

        wrapper.flush()
        wrapper.detach()

        return Response(
            output.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=libo_export.csv"},
        )

    elif fmt == "excel":
        if Workbook is None:
            return "openpyxl не установлен", 500

        wb = Workbook()
        ws_users = wb.active
        ws_users.title = "Пользователи"
        ws_users.append(["user_id", "username", "first_name", "last_name", "created_at", "quiz_count"])
        for u in users:
            ws_users.append([u["user_id"], u["username"], u["first_name"], u["last_name"], u["created_at"], u["quiz_count"]])

        ws_ratings = wb.create_sheet("Оценки")
        ws_ratings.append(["user_id", "cocktail_name", "rating", "review", "created_at"])
        for r in ratings:
            ws_ratings.append([r["user_id"], r["cocktail_name"], r["rating"], r["review"], r["created_at"]])

        ws_quizzes = wb.create_sheet("Подборы коктейлей")
        ws_quizzes.append(["id", "user_id", "alcoholic", "temperature", "taste", "tea_strength", "strength", "created_at"])
        for qs in quiz_sessions:
            ws_quizzes.append([qs["id"], qs["user_id"], qs["alcoholic"], qs["temperature"], qs["taste"], qs["tea_strength"], qs["strength"], qs["created_at"]])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=libo_export.xlsx"},
        )

    return "Неподдерживаемый формат", 400


if __name__ == "__main__":
    port = int(os.getenv("WEB_PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
